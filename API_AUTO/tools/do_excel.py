from openpyxl import load_workbook
from tools import project_path
from tools.read_config import ReadConfig
import json
from tools.get_data import GetData


class DoExcel:
    @classmethod
    def get_data(cls, file_name):
        wb = load_workbook(file_name)

        # print(sheet.cell(2, 1).value)
        mode = eval(ReadConfig.get_config(project_path.case_config_path, 'MODE', 'mode'))
        host = ReadConfig.get_config(project_path.case_config_path, 'MODE', 'host')


        test_data = []
        for key in mode:  # 遍历存放配置文件的字典
            sheet = wb[key]  # key 是表单名
            if mode[key] == 'all':

                for i in range(2, sheet.max_row + 1):  # max_row 最大行，  column 列
                    row_data = {}
                    row_data['case_id'] = sheet.cell(i, 1).value
                    row_data['url'] = sheet.cell(i, 2).value
                    if sheet.cell(i, 3).value:
                        row_data['data'] = json.loads(sheet.cell(i, 3).value)
                    else:
                        row_data['data'] = sheet.cell(i, 3).value
                    # row_data['data'] = sheet.cell(i, 3).value
                    # print(type(sheet.cell(i, 3).value))
                    # if sheet.cell(i, 3).value.find('${tel}') != -1:
                    #     row_data['data'] = sheet.cell(i, 3).value.replace('${tel}', tel)
                    # else:
                    #     row_data['data'] = sheet.cell(i, 3).value
                    row_data['title'] = sheet.cell(i, 4).value
                    row_data['http_method'] = sheet.cell(i, 5).value
                    row_data['expected'] = sheet.cell(i, 6).value  # 期望值
                    row_data['sheet_name'] = key
                    test_data.append(row_data)


            else:
                for case_id in mode[key]:
                    row_data = {}
                    row_data['case_id'] = sheet.cell(case_id + 1, 1).value
                    row_data['url'] = sheet.cell(case_id + 1, 2).value
                    if sheet.cell(case_id+1, 3).value:
                        row_data['data'] = json.loads(sheet.cell(case_id+1, 3).value)
                    else:
                        row_data['data'] = sheet.cell(case_id+1, 3).value

                    # row_data['data'] = sheet.cell(case_id + 1, 3).value
                    row_data['title'] = sheet.cell(case_id + 1, 4).value
                    row_data['http_method'] = sheet.cell(case_id + 1, 5).value
                    row_data['expected'] = sheet.cell(case_id + 1, 6).value  # 期望值
                    row_data['sheet_name'] = key
                    test_data.append(row_data)

        return test_data

    @staticmethod
    def write_back(file_name, sheet_name, i, result, TestResult):
        """用于专门写回数据"""
        wb = load_workbook(file_name)
        sheet = wb[sheet_name]
        sheet.cell(i, 7).value = result
        sheet.cell(i, 8).value = TestResult
        wb.save(file_name)  # 保存测试结果
        wb.close()

    @classmethod
    def update_tel(cls, tel, filename, sheet_name):
        wb = load_workbook(filename)
        sheet = wb[sheet_name]
        sheet.cell(2, 1).value = tel
        wb.save(filename)

    @classmethod
    def get_tel(cls, file_name, sheet_name):
        wb = load_workbook(file_name)
        sheet = wb[sheet_name]
        tel = sheet.cell(2, 1).value
        return tel


if __name__ == '__main__':
    test_data = DoExcel().get_data(project_path.test_case_path)
    print(test_data)
    # print(type(test_data))
    # print(test_data)
