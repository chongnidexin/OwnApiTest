# from tools.do_excel import DoExcel
# from tools import project_path
# import time
# from tools.get_user_info import GetUserInfo
#
# ts = int(time.time())
# sign = GetUserInfo().sign("*z5!2p(jywsjf-==!x_m+r4!lbs-#(gab4+nqwyi#n$9-4&7h_", ts, "15128283039")
# print(sign)
#
# # test_data = DoExcel.get_data(project_path.test_case_path)
#
# import requests
#
# req_body = {"tel": "15128283039", "code": "1234", "ts": ts, "sign": sign,
#             "client_public_key": ['''-----BEGIN PUBLIC KEY-----\nMIGfMA0GCSqGSIb3DQEBAQUAA4GNADCBiQKBgQCI3ZvndV2627UcC/hRrKCr0yS5\nbrPgODNpj+ExNzKMEfHqc1LG+iFaOcuRy4Ukt/UdsVN/EoZU4BxaDoFHxZxcnA9t\nIYQ0F6QQxztgouRQ9r1W6szYvtX69iFLXiRqPRDSqKjO0WKzwacfMNCeqL/2zyaU\n4pmDBJZrmjb18xTCWwIDAQAB\n-----END PUBLIC KEY-----'''],
#             "screen_width": 1080, "screen_height": 2016, "device_model": "16th", "device_identify": "60e71ae884988644",
#             "system_version": "8.1.0", "version": "5.1.2", "platform": "Android", "net_status": 1}
# url = "https://test.paquapp.com/user/login/4/"
# header = {
#     "Content-Type": "application/x-www-form-urlencoded",
#     "Content-Length": "555",
#     "Host": "test.paquapp.com",
#     "Connection": "Keep-Alive",
#     "Accept-Encoding": "gzip",
#     "User-Agent": "okhttp/3.12.1"
# }
# res = requests.post(url=url, data=req_body, headers=header)
# print(res.text)


import requests
from openpyxl import load_workbook

# 对应环境对应host
host = "http://boxonline-experience.paquapp.com"


def write_back(file_name, sheet_name, i, result, TestResult):
    """用于专门写回数据"""
    wb = load_workbook(file_name)
    sheet = wb[sheet_name]
    sheet.cell(i, 1).value = result
    sheet.cell(i, 2).value = TestResult
    wb.save(file_name)  # 保存测试结果
    wb.close()


# 用户open_id(来源为对应环境数据库player表)
open_ids = ["ohhQf5YBnWFoT7JxxUQW6En-IqqI", "oiNwG5rIt3v99ivc7mGphNYP4-Jk", "ohhQf5c291kz9RILb7LxusNVPbB0",
            "ohhQf5bCaxAPVNEygeBuEVds95Js", "ohhQf5VJqpamaPZECriojEnOlzYc", "oiNwG5r_8qWY_gFg2ab0Fpg4r7wQ",
            "ohhQf5RuGaSmJ1Soch7u4vW9mJDU", "ohhQf5ckzkKhXqrX5O5XrDnDWyD8", "ohhQf5fwJDcdmvlVyYwRQJyXhYe8",
            "ohhQf5WX9FVNRmRvy26ktqvuG-j0"]

req_body = {"v": "204"}
i = 1
# 登陆url
url = host + "/nf/logintest"
for open_id in open_ids:
    req_body['open_id'] = open_id
    res = requests.get(url=url, params=req_body)
    # print(res.json()['data']['token'])
    write_back('token.xlsx', 'Sheet1', i, 'token_' + str(i), res.json()['data']['token'])
    i += 1
